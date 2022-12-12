"""
============================
Project: py49Api
Author:柠檬班-海励
Time:2022/5/13 21:34
E-mail:2227092769@qq.com
Company:湖南零檬信息技术有限公司
Site: http://www.lemonban.com
Forum: http://testingpai.com 
============================
"""
from openpyxl import load_workbook


from tools.handle_path import data_dir

class HandleExcel:
    def __init__(self,file_name):
        self.wb = load_workbook(filename=file_name) #表
        # self.sheet = self.wb[sheet_name] # 表单
        #获取excel中所有的sheetname
        self.sheet_names = self.wb.sheetnames

    def __close_excel(self):
        self.wb.close()

    def get_data_cases(self,sheet_name):
        case_list = []
        sheet_obj = self.wb[sheet_name]
        data_list = list(sheet_obj.iter_rows(values_only=True))
        title = data_list[0]
        datas = data_list[1:]
        for case in datas:
            res = dict(zip(title, case))
            case_list.append(res)
        self.__close_excel()
        return case_list


    # def get_data(self):
    #     case_list = []
    #     data_list = list(self.sheet.iter_rows(values_only=True))
    #     title = data_list[0]
    #     datas = data_list[1:]
    #     for case in datas:
    #         res = dict(zip(title, case))
    #         case_list.append(res)
    #     self.__close_excel()
    #     return case_list

if __name__ == '__main__':
    cl = HandleExcel(file_name=data_dir,sheet_name="login")
    result = cl.get_data()[0]
    data = result["data"]
    print(type(data))
    import json
    import ast

    # 将excel中数据转换成python对象
    # res = json.loads(data)
    # res = eval(data)
    res = ast.literal_eval(data)
    print(type(res))
    print(res)


