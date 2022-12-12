"""
============================
Project: py49
Author:柠檬班-海励
Time:2022/4/13 20:57
E-mail:2227092769@qq.com
Company:湖南零檬信息技术有限公司
Site: http://www.lemonban.com
Forum: http://testingpai.com 
============================
"""

from openpyxl import load_workbook

# 获取表格对象
wb = load_workbook(filename="case_data.xlsx")

# 获取sheet对象
sheet = wb["login"]

# 写数据
# 覆盖写入
sheet["B2"] = "py49"
# # 覆盖写入 ，row:行, column：列, value=None：改数据的赋值
sheet.cell(row=2,column=2,value="登陆成功")
#追加写入，不会覆盖现有数据，在表格最后面写入数据，
sheet.append(["12","34","56"]) #接收可迭代对象


# 写行数据
for val in range(1,5):
    sheet.cell(row=1,column=val,value="登陆成功")

# 保存数据
wb.save(filename="case_data.xlsx")

wb.close()