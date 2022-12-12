"""
============================
Project: py49
Author:柠檬班-海励
Time:2022/4/13 20:38
E-mail:2227092769@qq.com
Company:湖南零檬信息技术有限公司
Site: http://www.lemonban.com
Forum: http://testingpai.com 
============================
"""
"""
六、行操作
1、行切片
result = sheet.iter_rows(min_row=1,max_row=2,min_col=1,max_col=4,values_only=True)
min_row=None: 行的起始索引值，默认是1，必须是int类型
max_row=None: 行的结束索引值，默认是1，必须是int类型
min_col=None: 列的起始索引值，默认是1，必须是int类型
max_col=None：列的结束索引值，默认是1，必须是int类型 
values_only=False：返回是对象
            True：返回对应的值

2、获取行和列的最大值
最大行
print(sheet.max_row)
最大列
print(sheet.max_column)

3、列切片
result = sheet.iter_cols(min_row=1,max_row=2,min_col=1,max_col=4,values_only=True)
min_row=None: 行的起始索引值，默认是1，必须是int类型
max_row=None: 行的结束索引值，默认是1，必须是int类型
min_col=None: 列的起始索引值，默认是1，必须是int类型
max_col=None：列的结束索引值，默认是1，必须是int类型 
values_only=False：返回是对象
            True：返回对应的值
"""

from openpyxl import load_workbook

# 获取表格对象
wb = load_workbook(filename="case_data.xlsx")

# 获取sheet对象
sheet = wb["login"]

result = sheet.iter_rows(values_only=True)
# result = sheet.iter_rows(min_row=1,max_row=2,min_col=1,max_col=4,values_only=True)
print(list(result))
# for i in list(result):
#     for ii in i:
#         print(ii.value)


#用不着，竖着读数据
# result = sheet.iter_cols(min_row=1,max_row=2,min_col=1,max_col=4,values_only=True)
# print(list(result))


wb.close()



