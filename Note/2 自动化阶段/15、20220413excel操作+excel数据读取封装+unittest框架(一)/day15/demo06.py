"""
============================
Project: py49
Author:柠檬班-海励
Time:2022/4/13 21:26
E-mail:2227092769@qq.com
Company:湖南零檬信息技术有限公司
Site: http://www.lemonban.com
Forum: http://testingpai.com 
============================
"""

"""
# 获取表格对象
wb = load_workbook(filename="case_data.xlsx")
case_list = []
# 获取sheet对象
sheet = wb["login"]
data_list  = list(sheet.iter_rows(values_only=True))
# zip函数 ，数据标题，数据行
title = data_list[0]
datas = data_list[1:]
print(title)
print(data_list)
for case in datas:
    res = dict(zip(title,case))
    case_list.append(res)
pprint.pprint(case_list)
print(case_list)
wb.close()


[(id,title,url ),(1,登陆成功),(2，登陆失败)]
(id,title,url )
[(1,登陆成功),(2，登陆失败)]
[(id,1),(title,登陆成功)] 通过zip函数 就得到字典{id:1,title:登陆成功}
"""
from openpyxl import load_workbook
class HandleExcel:
    def __init__(self,file_name,sheet_name):
        self.wb = load_workbook(filename=file_name) #表
        self.sheet = self.wb[sheet_name] # 表单

    def __close_excel(self):
        self.wb.close()

    def get_data(self):
        case_list = []
        data_list = list(self.sheet.iter_rows(values_only=True))
        title = data_list[0]
        datas = data_list[1:]
        for case in datas:
            res = dict(zip(title, case))
            case_list.append(res)
        self.__close_excel()
        return case_list

if __name__ == '__main__':
    cl = HandleExcel(file_name="case_data.xlsx",sheet_name="login")
    result = cl.get_data()
    print(result)




