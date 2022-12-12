"""
============================
Project: py49
Author:柠檬班-海励
Time:2022/4/13 20:10
E-mail:2227092769@qq.com
Company:湖南零檬信息技术有限公司
Site: http://www.lemonban.com
Forum: http://testingpai.com 
============================
"""
"""
一、excel操作
1、存放测试用例数据，数据驱动的数据来源

二、安装
pip install openpyxl -i http://mirrors.aliyun.com/pypi/simple/

[{"id":"1","title":"登陆成功"},{"id":"1","title":"登陆失败"}]

三、表结构
1、表  ==》 对象
2、sheet==》 对象
3、单元格 ==》 对象

{表:{sheet1:{单元格1:"val1"，单元格2:"val2"}}}

四、操作步骤
from openpyxl import load_workbook
# 获取表格对象
wb = load_workbook(filename="case_data.xlsx")
# 获取sheet对象
sheet = wb["login"]
# 获取单元格对象
cell = sheet["B2"]
# 获取值
print(cell.value)
# 关闭数据流
wb.close()

五、load_workbook参数
1、filename：文件名称
2、read_only：只读模式
3、data_only：读取数据的时候如果遇到计算公式是否要计算后再读取数据，
              False:读取未计算的结果(公式读出来) True：读取计算后的记过
              
              

"""


from openpyxl import load_workbook

# 获取表格对象
wb = load_workbook(filename="case_data.xlsx")

# 获取sheet对象
# sheet = wb["login"]
# 获取所有的sheet名称
# sheets = wb.sheetnames
# print(sheets)
# 通过索引获取sheet
sheet = wb.worksheets[0]

# 获取单元格对象
# cell = sheet["B2"]
# row:行, column：列, value=None：改数据的赋值
cell = sheet.cell(row=2,column=2)

# 获取值
print(cell.value)

wb.close()




