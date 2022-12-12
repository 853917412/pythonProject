"""
============================
Project: py49
Author:柠檬班-海励
Time:2022/4/13 21:15
E-mail:2227092769@qq.com
Company:湖南零檬信息技术有限公司
Site: http://www.lemonban.com
Forum: http://testingpai.com 
============================
"""
"""

[{"id":"1","title":"登陆成功"},{"id":"1","title":"登陆失败"}]

"""

import pprint
from openpyxl import load_workbook

# 获取表格对象
wb = load_workbook(filename="case_data.xlsx")
case_list = []
# 获取sheet对象
sheet = wb["login"]
data_list  = list(sheet.iter_rows(values_only=True))
# zip函数 ，数据标题，数据行
title = data_list[0]
datas = data_list[1:]
print(title)
print(data_list)
for case in datas:
    res = dict(zip(title,case))
    case_list.append(res)
pprint.pprint(case_list)
print(case_list)
wb.close()