from openpyxl import load_workbook





if __name__ == "__main__":
    #获取表对象
    wb = load_workbook(filename=(r'D:\pythonProject\Note\2 自动化阶段\15、20220413excel操作+excel数据读取封装+unittest框架(一)\day15\case_data.xlsx'))
    #获取sheet对象
    sheet_obj = wb['login']
    #行切片
    result = sheet_obj.iter_rows(min_row=1,max_row=1,min_col=3,max_col=3,values_only=True)
    print(sheet_obj.max_column)
    wb.close()




